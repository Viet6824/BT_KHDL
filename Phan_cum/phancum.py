"""
Phân cụm sinh viên lớp K58KTP bằng thuật toán K-Means (K=3)
Tự động gom nhóm dựa trên sự phân bố điểm số thực tế của sinh viên.
GPA được làm tròn chuẩn xác đến 2 chữ số thập phân.
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import rcParams
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from sklearn.cluster import KMeans
import warnings
warnings.filterwarnings("ignore")

rcParams["font.family"] = "DejaVu Sans"

# ============================================================
# 1. ĐỌC & LÀM SẠCH DỮ LIỆU
# ============================================================
df_raw       = pd.read_excel(r"D:\Phan_cum\TỔNG HỢP ĐIỂM K58KTP.xlsx", header=None)
mssv         = df_raw.iloc[1, 3:].values
names        = df_raw.iloc[2, 3:].values
subject_names = df_raw.iloc[4:, 2].values   # tên 52 môn
scores_raw   = df_raw.iloc[4:, 3:]

def clean_score(val):
    try:
        v = float(str(val).replace(",", ".").strip())
        return v if 0 <= v <= 4.0 else np.nan
    except Exception:
        return np.nan

scores_clean = scores_raw.apply(lambda col: col.map(clean_score))
scores_clean = scores_clean.reset_index(drop=True)
scores_clean.columns = range(scores_clean.shape[1])
X_raw = scores_clean.T.reset_index(drop=True)  # (n_sv, n_mon)

# ============================================================
# 2. PHÂN CỤM DỮ LIỆU BẰNG THUẬT TOÁN K-MEANS
# ============================================================
gpa = X_raw.mean(axis=1, skipna=True)
valid_mask = gpa.notna() & (X_raw.notna().sum(axis=1) >= 5)

result_rows = []
for i in range(len(names)):
    if not valid_mask.iloc[i]:
        continue
    result_rows.append({
        "MSSV":          mssv[i],
        "Họ và tên":     names[i],
        "GPA TB":        round(gpa.iloc[i], 2),  
        "_scores":       X_raw.iloc[i].values,
    })

result_df = pd.DataFrame(result_rows)

# Thực hiện thuật toán K-Means Clustering với K = 3 cụm
X_kmeans = result_df[["GPA TB"]].values
kmeans = KMeans(n_clusters=3, random_state=42, n_init=10)
result_df["Cluster_Raw"] = kmeans.fit_predict(X_kmeans)

# Định hướng lại nhãn cụm theo thứ tự điểm số giảm dần
avg_gpa_per_cluster = result_df.groupby("Cluster_Raw")["GPA TB"].mean().sort_values(ascending=False)
cluster_mapping = {old_label: new_label for new_label, old_label in enumerate(avg_gpa_per_cluster.index)}
result_df["Cluster"] = result_df["Cluster_Raw"].map(cluster_mapping)

# Đặt tên trực quan cho các cụm
cluster_names = {0: "Nhóm điểm cao", 1: "Nhóm điểm trung bình", 2: "Nhóm điểm thấp"}
result_df["Xếp loại"] = result_df["Cluster"].map(cluster_names)

# Sắp xếp danh sách đầu ra
result_df = result_df.sort_values(["Cluster", "GPA TB"], ascending=[True, False]).reset_index(drop=True)

# Thống kê số lượng sinh viên mỗi cụm
LABELS = ["Nhóm điểm cao", "Nhóm điểm trung bình", "Nhóm điểm thấp"]
counts = result_df["Xếp loại"].value_counts().reindex(LABELS, fill_value=0)

print("Kết quả phân cụm học máy K-Means:")
for k, v in counts.items():
    grp = result_df[result_df["Xếp loại"] == k]
    print(f"  {k}: {v} SV  |  GPA TB của nhóm: {grp['GPA TB'].mean():.2f}") 

# ============================================================
# 3. VẼ BIỂU ĐỒ CỘT PHÂN CỤM
# ============================================================
COLORS   = ["#27AE60", "#E67E22", "#E74C3C"]
vals     = [counts["Nhóm điểm cao"], counts["Nhóm điểm trung bình"], counts["Nhóm điểm thấp"]]
gpas     = [result_df[result_df["Xếp loại"] == k]["GPA TB"].mean() for k in LABELS]

fig, ax = plt.subplots(figsize=(10, 7), facecolor="#F8F9FA")
ax.set_facecolor("#F4F6F7")

display_labels = []
for nhom in LABELS:
    grp = result_df[result_df["Xếp loại"] == nhom]
    if len(grp) > 0:
        display_labels.append(f"{nhom}\n({grp['GPA TB'].min():.2f} ≤ GPA ≤ {grp['GPA TB'].max():.2f})")
    else:
        display_labels.append(f"{nhom}\n(Không có SV)")

bars = ax.bar(display_labels, vals, color=COLORS, width=0.5, edgecolor="white", linewidth=2, zorder=3)

for bar, cnt, gpa_v in zip(bars, vals, gpas):
    if cnt > 0:
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                f"{cnt} sinh viên", ha="center", va="bottom", fontsize=13, fontweight="bold", color="#2C3E50")
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() / 2,
                f"GPA TB Nhóm\n{gpa_v:.2f}", ha="center", va="center", fontsize=11, color="white", fontweight="bold") 

ax.set_title("BIỂU ĐỒ PHÂN CỤM HỌC LỰC K-MEANS – LỚP K58KTP", fontsize=16, fontweight="bold", color="#1A252F", pad=20)
ax.set_ylabel("Số sinh viên", fontsize=12, color="#555")
ax.set_ylim(0, max(vals) * 1.25 if max(vals) > 0 else 10)
ax.grid(axis="y", alpha=0.4, linestyle="--", zorder=0)
ax.spines[["top", "right", "left"]].set_visible(False)
ax.tick_params(axis="x", labelsize=10)

total = sum(vals)
if total > 0:
    fig.text(0.99, 0.01,
             f"Tổng: {total} SV  |  Nhóm cao: {vals[0]/total*100:.1f}%  "
             f"Nhóm TB: {vals[1]/total*100:.1f}%  Nhóm thấp: {vals[2]/total*100:.1f}%",
             ha="right", va="bottom", fontsize=9, color="#999", style="italic")

chart_path = "phan_cum_K58KTP.png"
plt.tight_layout()
plt.savefig(chart_path, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
print(f"Đã lưu biểu đồ phân cụm: {chart_path}")

# ============================================================
# 4. XUẤT EXCEL BÁO CÁO KẾT QUẢ
# ============================================================
def thin_border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_al():
    return Alignment(horizontal="left", vertical="center")

GRP_BG   = {"Nhóm điểm cao": "D5F5E3", "Nhóm điểm trung bình": "FDEBD0", "Nhóm điểm thấp": "FADBD8"}
GRP_FONT = {"Nhóm điểm cao": "1E8449", "Nhóm điểm trung bình": "935116", "Nhóm điểm thấp": "922B21"}
GRP_HDR  = {"Nhóm điểm cao": "27AE60", "Nhóm điểm trung bình": "E67E22", "Nhóm điểm thấp": "E74C3C"}

wb = Workbook()

# ═══════════════════════════════════════════════════════════
# Sheet 1: Tổng quan phân cụm
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Tổng quan"
ws1.sheet_view.showGridLines = False

ws1.merge_cells("A1:G2")
c = ws1["A1"]
c.value     = "KẾT QUẢ PHÂN CỤM HỌC LỰC BẰNG K-MEANS – LỚP K58KTP"
c.font      = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
c.fill      = PatternFill("solid", start_color="1A5276")
c.alignment = center()
ws1.row_dimensions[1].height = 22
ws1.row_dimensions[2].height = 22

ws1.merge_cells("A3:G3")
c2 = ws1["A3"]
c2.value     = "Phương pháp phân tích: Định cụm tự động bằng thuật toán học máy K-Means Clustering (K=3)"
c2.font      = Font(name="Calibri", italic=True, size=11, color="5D6D7E")
c2.fill      = PatternFill("solid", start_color="EBF5FB")
c2.alignment = center()
ws1.row_dimensions[3].height = 18

stat_hdrs = ["Tên Cụm Hệ Thống", "Ranh giới GPA thực tế", "Số SV", "GPA TB Cụm", "GPA Cao nhất", "GPA Thấp nhất", "Tỉ lệ"]
for col, h in enumerate(stat_hdrs, 1):
    c = ws1.cell(row=5, column=col, value=h)
    c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color="1A5276")
    c.alignment = center()
    c.border    = thin_border()
ws1.row_dimensions[5].height = 22

criteria = {}
for nhom in LABELS:
    grp = result_df[result_df["Xếp loại"] == nhom]
    if len(grp) > 0:
        criteria[nhom] = f"{grp['GPA TB'].min():.2f} → {grp['GPA TB'].max():.2f}"
    else:
        criteria[nhom] = "-"

for row_i, nhom in enumerate(LABELS, 6):
    grp = result_df[result_df["Xếp loại"] == nhom]
    n   = len(grp)
    vals_row = [
        nhom,
        criteria[nhom],
        n,
        round(grp["GPA TB"].mean(), 2) if n else "-",  
        round(grp["GPA TB"].max(), 2)  if n else "-",   
        round(grp["GPA TB"].min(), 2)  if n else "-",   
        f"{n/total*100:.1f}%" if total > 0 else "0%",
    ]
    for col, v in enumerate(vals_row, 1):
        c = ws1.cell(row=row_i, column=col, value=v)
        c.fill      = PatternFill("solid", start_color=GRP_BG[nhom])
        c.font      = Font(name="Calibri", size=11, bold=(col == 1), color=GRP_FONT[nhom])
        c.alignment = center() if col != 1 else left_al()
        c.border    = thin_border()
    ws1.row_dimensions[row_i].height = 20

col_ws1 = [28, 22, 8, 14, 14, 14, 10]
for i, w in enumerate(col_ws1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

try:
    img = XLImage(chart_path)
    img.width, img.height = 700, 490
    ws1.add_image(img, "A10")
except Exception as e:
    print(f"Không thể nhúng ảnh biểu đồ vào Excel: {e}")

# ═══════════════════════════════════════════════════════════
# Sheet 2: Danh sách chi tiết kèm nhãn cụm K-Means
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Danh sách chi tiết")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "D3"

n_cols_total = 3 + len(subject_names) + 2
ws2.merge_cells(f"A1:{get_column_letter(n_cols_total)}1")
c = ws2["A1"]
c.value     = "DANH SÁCH CHI TIẾT ĐIỂM SỐ VÀ NHÃN PHÂN CỤM THEO THUẬT TOÁN K-MEANS"
c.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
c.fill      = PatternFill("solid", start_color="1A5276")
c.alignment = center()
ws2.row_dimensions[1].height = 28

fixed_hdrs = ["STT", "MSSV", "Họ và tên"]
tail_hdrs  = ["GPA TB", "Kết quả cụm"]
all_hdrs   = fixed_hdrs + list(subject_names) + tail_hdrs

for col, h in enumerate(all_hdrs, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font      = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    c.alignment = center(wrap=True)
    c.border    = thin_border()
    if col <= 3:
        c.fill = PatternFill("solid", start_color="1A5276")
    elif col <= 3 + len(subject_names):
        c.fill = PatternFill("solid", start_color="2E86C1")
    else:
        c.fill = PatternFill("solid", start_color="1E8449")
ws2.row_dimensions[2].height = 55

for r_i, row in result_df.iterrows():
    nhom  = row["Xếp loại"]
    rrow  = r_i + 3
    alt   = "F2F3F4" if r_i % 2 == 0 else "FFFFFF"

    for col, v in enumerate([r_i+1, row["MSSV"], row["Họ và tên"]], 1):
        c = ws2.cell(row=rrow, column=col, value=v)
        c.fill      = PatternFill("solid", start_color=alt)
        c.font      = Font(name="Calibri", size=9, bold=(col==3))
        c.alignment = left_al() if col == 3 else center()
        c.border    = thin_border()

    scores_i = row["_scores"]
    for j, score_val in enumerate(scores_i):
        col = j + 4
        v   = round(float(score_val), 1) if not np.isnan(score_val) else ""
        c   = ws2.cell(row=rrow, column=col, value=v)
        c.font      = Font(name="Calibri", size=9)
        c.alignment = center()
        c.border    = thin_border()
        if v != "" and isinstance(v, float):
            if v >= 3.5:
                c.fill = PatternFill("solid", start_color="D5F5E3")
                c.font = Font(name="Calibri", size=9, color="1E8449")
            else:
                c.fill = PatternFill("solid", start_color=alt)
        else:
            c.fill = PatternFill("solid", start_color="F8F9FA")

    tail_vals = [row["GPA TB"], nhom]
    for j, v in enumerate(tail_vals):
        col = 3 + len(subject_names) + 1 + j
        c   = ws2.cell(row=rrow, column=col, value=v)
        c.border    = thin_border()
        c.alignment = center()
        c.fill = PatternFill("solid", start_color=GRP_BG[nhom])
        c.font = Font(name="Calibri", size=9, bold=True, color=GRP_FONT[nhom])

    ws2.row_dimensions[rrow].height = 16

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 22
for j in range(len(subject_names)):
    ws2.column_dimensions[get_column_letter(j + 4)].width = 8
last2_start = 3 + len(subject_names) + 1
ws2.column_dimensions[get_column_letter(last2_start)].width     = 10
ws2.column_dimensions[get_column_letter(last2_start + 1)].width = 18

# ═══════════════════════════════════════════════════════════
# Sheet 3: Phân tách riêng từng cụm K-Means
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Phân nhóm")
ws3.sheet_view.showGridLines = False

cur = 1
for nhom in LABELS:
    grp = result_df[result_df["Xếp loại"] == nhom].reset_index(drop=True)

    ws3.merge_cells(f"A{cur}:E{cur}") 
    
    # ĐÃ ĐỔI & SỬA LỖI: Tách biến riêng để tránh lỗi cú pháp f-string
    mean_gpa = grp['GPA TB'].mean() if len(grp) > 0 else 0
    
    c = ws3.cell(row=cur, column=1,
                 value=f"  {nhom.upper()}  —  {len(grp)} SV  |  GPA TB Cụm: {mean_gpa:.2f}  |  Khoảng điểm: {criteria[nhom]}") 
    c.font      = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=GRP_HDR[nhom])
    c.alignment = left_al()
    ws3.row_dimensions[cur].height = 24
    cur += 1

    for col, h in enumerate(["STT", "MSSV", "Họ và tên", "GPA Tích Lũy", "Nhãn Cụm"], 1):
        c = ws3.cell(row=cur, column=col, value=h)
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color="2C3E50")
        c.alignment = center()
        c.border    = thin_border()
    ws3.row_dimensions[cur].height = 18
    cur += 1

    for i, row in grp.iterrows():
        alt = GRP_BG[nhom] if i % 2 == 0 else "FFFFFF"
        for col, v in enumerate([i+1, row["MSSV"], row["Họ và tên"], row["GPA TB"], row["Xếp loại"]], 1):
            c = ws3.cell(row=cur, column=col, value=v)
            c.fill      = PatternFill("solid", start_color=alt)
            c.border    = thin_border()
            c.font      = Font(name="Calibri", size=10, bold=(col in [4, 5]),
                               color=GRP_FONT[nhom] if col == 5 else "2C3E50")
            c.alignment = left_al() if col == 3 else center()
        ws3.row_dimensions[cur].height = 17
        cur += 1
    cur += 1

for i, w in enumerate([5, 16, 26, 15, 18], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

xl_path = "ket_qua_phan_cum.xlsx"
wb.save(xl_path)
print(f"Đã lưu kết quả Excel: {xl_path}")
print("\n✅ HOÀN THÀNH QUY TRÌNH PHÂN CỤM HỌC MÁY K-MEANS!")